from openpyxl import load_workbook


class excel2sql():

    def extract_excel_data(self, filePath):

        wb = load_workbook(filename=filePath)
        print(wb.sheetnames)
        toolSheet = wb[wb.sheetnames[0]]

        # for row in toolSheet.iter_rows(min_row=1, max_col=4, max_row=4):
        data = []
        value_list = []
        for row in toolSheet.iter_rows(min_row=2):
            row_data = []
            values = '('
            for cell in row:
                print(cell.value)
                if str(cell.value).isnumeric():
                    print("number detected")
                    values = values + "{},".format(cell.value)
                else:
                    print("non number")
                    values = values + " '{}',".format(cell.value)
                row_data.append(cell.value)
            data.append(row_data)
            values = values[:-1]
            values = values + ")"
            value_list.append(values)

        print(value_list)
        return value_list
